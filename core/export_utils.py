"""
Export Utilities for SmartSpace UPY Admin
Provides Excel and PDF export functionality for reports
"""
from io import BytesIO
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_excel_response(filename: str) -> tuple:
    """Create Excel workbook and response object"""
    wb = Workbook()
    ws = wb.active
    return wb, ws, filename


def style_excel_header(ws, headers: list, row: int = 1):
    """Apply styling to Excel header row"""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = max(len(str(header)) + 5, 15)


def finalize_excel_response(wb, filename: str) -> HttpResponse:
    """Save workbook to response"""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_users_excel(users_queryset) -> HttpResponse:
    """Generate Excel file for Users list"""
    wb, ws, _ = create_excel_response("users.xlsx")
    ws.title = "Daftar Users"
    
    # Headers
    headers = ["No", "NPM/NIP", "Nama Lengkap", "Email", "Fakultas", "Program Studi", "Angkatan", "Role", "Status", "Tanggal Daftar"]
    style_excel_header(ws, headers)
    
    # Data
    for idx, user in enumerate(users_queryset, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=user.npm_nip or "-")
        ws.cell(row=row, column=3, value=user.get_full_name() or user.username)
        ws.cell(row=row, column=4, value=user.email or "-")
        ws.cell(row=row, column=5, value=user.get_fakultas_display() if user.fakultas else "-")
        ws.cell(row=row, column=6, value=user.program_studi or "-")
        ws.cell(row=row, column=7, value=user.angkatan or "-")
        ws.cell(row=row, column=8, value=user.role or "-")
        ws.cell(row=row, column=9, value="Aktif" if user.is_active else "Nonaktif")
        ws.cell(row=row, column=10, value=timezone.localtime(user.date_joined).strftime("%d/%m/%Y %H:%M"))
    
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    return finalize_excel_response(wb, f"users_{timestamp}.xlsx")


def generate_bookings_excel(bookings_queryset) -> HttpResponse:
    """Generate Excel file for Bookings list"""
    wb, ws, _ = create_excel_response("peminjaman.xlsx")
    ws.title = "Daftar Peminjaman"
    
    # Headers
    headers = ["No", "ID", "NPM/NIP", "Nama Peminjam", "Ruangan", "Tanggal Mulai", "Tanggal Selesai", "Jumlah Tamu", "Status", "Dibuat"]
    style_excel_header(ws, headers)
    
    # Data
    for idx, booking in enumerate(bookings_queryset, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=booking.id)
        ws.cell(row=row, column=3, value=booking.user.npm_nip or "-")
        ws.cell(row=row, column=4, value=booking.user.get_full_name() or booking.user.username)
        ws.cell(row=row, column=5, value=booking.room.nomor_ruangan)
        ws.cell(row=row, column=6, value=timezone.localtime(booking.tanggal_mulai).strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row, column=7, value=timezone.localtime(booking.tanggal_selesai).strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row, column=8, value=booking.jumlah_tamu or 1)
        ws.cell(row=row, column=9, value=booking.get_status_display())
        ws.cell(row=row, column=10, value=timezone.localtime(booking.created_at).strftime("%d/%m/%Y %H:%M"))
    
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    return finalize_excel_response(wb, f"peminjaman_{timestamp}.xlsx")


def generate_dashboard_excel(stats: dict) -> HttpResponse:
    """Generate Excel file for Dashboard Report"""
    wb, ws, _ = create_excel_response("dashboard_report.xlsx")
    ws.title = "Laporan Dashboard"
    
    # Title
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = "LAPORAN DASHBOARD SMARTSPACE UPY"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    ws.cell(row=2, column=1, value=f"Digenerate: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
    
    # Summary Section
    ws.cell(row=4, column=1, value="RINGKASAN").font = Font(bold=True)
    summary_data = [
        ("Total Booking", stats['summary']['total']),
        ("Pending", stats['summary']['pending']),
        ("Approved", stats['summary']['approved']),
        ("Ruangan Aktif", stats['summary']['active_rooms']),
    ]
    
    for idx, (label, value) in enumerate(summary_data, 5):
        ws.cell(row=idx, column=1, value=label)
        ws.cell(row=idx, column=2, value=value)
    
    # Status Distribution
    start_row = 10
    ws.cell(row=start_row, column=1, value="DISTRIBUSI STATUS").font = Font(bold=True)
    style_excel_header(ws, ["Status", "Jumlah"], start_row + 1)
    
    for idx, item in enumerate(stats['status_distribution'], start_row + 2):
        ws.cell(row=idx, column=1, value=item['status'])
        ws.cell(row=idx, column=2, value=item['count'])
    
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    return finalize_excel_response(wb, f"dashboard_report_{timestamp}.xlsx")


# PDF Export Functions
def generate_bookings_pdf(bookings_queryset) -> HttpResponse:
    """Generate PDF file for Bookings Report"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=1, spaceAfter=20)
    elements.append(Paragraph("LAPORAN PEMINJAMAN RUANGAN", title_style))
    elements.append(Paragraph(f"SmartSpace UPY - {timezone.now().strftime('%d %B %Y')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table Data
    data = [["No", "Peminjam", "Ruangan", "Tanggal", "Waktu", "Status"]]
    
    for idx, booking in enumerate(bookings_queryset[:100], 1):  # Limit to 100 for PDF
        tanggal = timezone.localtime(booking.tanggal_mulai).strftime("%d/%m/%Y")
        waktu = f"{timezone.localtime(booking.tanggal_mulai).strftime('%H:%M')} - {timezone.localtime(booking.tanggal_selesai).strftime('%H:%M')}"
        data.append([
            str(idx),
            booking.user.get_full_name() or booking.user.username,
            booking.room.nomor_ruangan[:20],
            tanggal,
            waktu,
            booking.get_status_display()
        ])
    
    table = Table(data, colWidths=[0.4*inch, 2*inch, 2*inch, 1.2*inch, 1.5*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="peminjaman_{timestamp}.pdf"'
    return response


def generate_dashboard_pdf(stats: dict) -> HttpResponse:
    """Generate PDF file for Dashboard Report"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import inch
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=1, spaceAfter=20)
    elements.append(Paragraph("LAPORAN DASHBOARD", title_style))
    elements.append(Paragraph(f"SmartSpace UPY - {timezone.now().strftime('%d %B %Y %H:%M')}", 
                              ParagraphStyle('Subtitle', parent=styles['Normal'], alignment=1)))
    elements.append(Spacer(1, 30))
    
    # Summary
    elements.append(Paragraph("Ringkasan Statistik", styles['Heading2']))
    elements.append(Spacer(1, 10))
    
    summary_data = [
        ["Metrik", "Nilai"],
        ["Total Booking", str(stats['summary']['total'])],
        ["Pending", str(stats['summary']['pending'])],
        ["Approved", str(stats['summary']['approved'])],
        ["Ruangan Aktif", str(stats['summary']['active_rooms'])],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Status Distribution
    elements.append(Paragraph("Distribusi Status", styles['Heading2']))
    elements.append(Spacer(1, 10))
    
    status_data = [["Status", "Jumlah"]]
    for item in stats['status_distribution']:
        status_data.append([item['status'], str(item['count'])])
    
    status_table = Table(status_data, colWidths=[3*inch, 2*inch])
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(status_table)
    
    doc.build(elements)
    
    buffer.seek(0)
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="dashboard_report_{timestamp}.pdf"'
    return response
